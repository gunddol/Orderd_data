from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl import load_workbook
import tkinter.messagebox

global data_count
yes_data = [0] * 11

# 자료 종류 및 수량 추합 -> 콤보박스 체크
def data_sort(data):
    cnt = -1
    data_count = 0
    for i in combo_list:
        cnt = cnt + 1
        if i.get() != "":
            yes_data[data_count] = cnt
            data_count = data_count + 1

    global user_data_list
    global user_num
    user_data_list = [""] * 5000
    user_num = 2
    row_index = "A"+ str(user_num)
    while (data[row_index].value != None and data["B" + str(user_num)].value != None and data["C" + str(user_num)].value != None):
        user_data_dic = {}
        for i in range(data_count):
            tmp = combo_list[yes_data[i]].get()
            row_index = str(tmp[0]) + str(user_num)
            user_data_dic[info_list[yes_data[i]]["text"]] = data[row_index].value

        user_data_list[user_num-1] = user_data_dic
        user_num = user_num + 1
        row_index = "A" + str(user_num)

    print(user_data_list)
    #user_num = 2 + 구매자 수 + 1
    # print(user_data_list[3]['주소'])

# 자료정리
def sheet_import():
    try:
        load_sheet = load_workbook(excel_name,data_only=True)['Sheet1']
        data_sort(load_sheet)
        check_accepted.config(text = "성공!")
    except :
        check_accepted.config(text = "실패ㅠ")
        tkinter.messagebox.showwarning("파일 불러오기 오류","시트의 이름을 확인하세요 '\n' 시트 이름 = 'Sheet1'으로 하십시오.")


# 엑셀 파일 불러오기
def load_excel():
    window.filename = filedialog.askopenfilename(initialdir="C:/User/", title="choose your file", filetypes=(("Excel files", "*.xls"), ("Excel files", "*.xlsx"), ("all files", "*.*")))
    global excel_name
    excel_name = str(window.filename)
    print(excel_name)
    root = window.filename.split("/")
    label2.config(text = root[len(root)-1])


def load_goal():
    window.filename = filedialog.askopenfilename(initialdir="C:/User/", title="choose your file", filetypes=(("Excel files", "*.xls"), ("Excel files", "*.xlsx"), ("all files", "*.*")))
    global excel_name2
    excel_name2 = str(window.filename)
    print(excel_name2)
    root = window.filename.split("/")
    label3.config(text=root[len(root) - 1])

def sheet_export():
    try:
        load_file = load_workbook(excel_name2, data_only=True)
        load_sheet = load_file.get_sheet_by_name('택배출고요청 발주서')

        for i in range(1, user_num - 1):
            row_num = str(int(start_row.get()) + i - 1)
            load_sheet['A' + row_num] = user_data_list[i]['받는분']

            if '주연락처' in user_data_list[i]:
                load_sheet['B' + row_num] = user_data_list[i]['주연락처']

            if '부연락처' in user_data_list[i]:
                load_sheet['C' + row_num] = user_data_list[i]['부연락처']

            load_sheet['E' + row_num] = user_data_list[i]['주소']
            load_sheet['I' + row_num] = user_data_list[i]['구매수량']
            if '배송메세지' in user_data_list[i]:
               load_sheet['J' + row_num] = user_data_list[i]['배송메세지']

            if user_data_list[i]['상품명'].find('닭') > 0:
                load_sheet['H' + row_num] = '위펫영양닭죽(10팩)'

            if user_data_list[i]['상품명'].find('단호') > 0:
                load_sheet['H' + row_num] = '위펫오리단호박죽(10팩)'

            if user_data_list[i]['상품명'].find('황태') > 0:
                load_sheet['H' + row_num] = '위펫영양황태죽(10팩)'

        load_file.save(excel_name2)

        check_accepted2.config(text = "성공!")
    except:
        check_accepted2.config(text = "실패ㅠ")
        tkinter.messagebox.showwarning("파일 저장하기 오류","시트의 이름을 확인하세요 '\n' 시트 이름 = '택배출고요청 발주서'으로 하십시오.")



window = Tk()
window.title("발주 데이터 양식 Coverter")
window.geometry("950x300")
window.resizable(False, False)

label1 = ttk.Label(window, text="발주서",anchor="center")
# print(label1["text"])
label1.place(x=10,y=10)

button1 = ttk.Button(window,text="불러오기", command = lambda:load_excel())
button1.place(x=60,y=8)

label2 = ttk.Label(window, text="파일 이름",width=100)
label2.place(x=200,y=10)

label_space = ttk.Label(window, text="",width=100)
label_space.place(x=10, y=40)


read_frame = ttk.Frame(window)

cell_column = [(chr(i)) +"열" for i in range(65, 80)]

info_1 = ttk.Label(read_frame, text="발송일자",width=10,anchor="center")
info_1.grid(row=2,column=0,columnspan=1)
combo_1 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_1.grid(row=3,column=0)

info_2 = ttk.Label(read_frame, text="주문자",width=10,anchor="center")
info_2.grid(row=2,column=1)
combo_2 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_2.grid(row=3,column=1)

info_3 = ttk.Label(read_frame, text="받는분",width=10,anchor="center")
info_3.grid(row=2,column=2)
combo_3 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_3.grid(row=3,column=2)

info_4 = ttk.Label(read_frame, text="주연락처",width=10,anchor="center")
info_4.grid(row=2,column=3)
combo_4 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_4.grid(row=3,column=3)

info_5 = ttk.Label(read_frame, text="부연락처",width=10,anchor="center")
info_5.grid(row=2,column=4)
combo_5 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_5.grid(row=3,column=4)

info_6 = ttk.Label(read_frame, text="주소",width=10,anchor="center")
info_6.grid(row=2,column=5)
combo_6 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_6.grid(row=3,column=5)

info_7 = ttk.Label(read_frame, text="상품코드",width=10,anchor="center")
info_7.grid(row=2,column=6)
combo_7 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_7.grid(row=3,column=6)

info_8 = ttk.Label(read_frame, text="상품명",width=10,anchor="center")
info_8.grid(row=2,column=7)
combo_8 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_8.grid(row=3,column=7)

info_9 = ttk.Label(read_frame, text="구매수량",width=10,anchor="center")
info_9.grid(row=2,column=8)
combo_9 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_9.grid(row=3,column=8)

info_10 = ttk.Label(read_frame, text="운송장번호",width=10,anchor="center")
info_10.grid(row=2,column=9)
combo_10 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_10.grid(row=3,column=9)

info_11 = ttk.Label(read_frame, text="배송메세지",width=10,anchor="center")
info_11.grid(row=2,column=10)
combo_11 = ttk.Combobox(read_frame, values=cell_column, width=5)
combo_11.grid(row=3,column=10)

button2 = ttk.Button(read_frame,text="자료가져오기", command = lambda:sheet_import())
button2.grid(row=3,column=11)

check_accepted = ttk.Label(read_frame, text="")
check_accepted.grid(row=3,column=12)

read_frame.place(x=10,y=50)

global info_list
info_list = [info_1, info_2, info_3, info_4, info_5, info_6, info_7, info_8, info_9, info_10, info_11]
global combo_list
combo_list = [combo_1, combo_2, combo_3, combo_4, combo_5, combo_6, combo_7, combo_8, combo_9, combo_10, combo_11]




button3 = ttk.Button(window,text="저장 할 파일", command = lambda:load_goal())
button3.place(x=10,y=148)

label3 = ttk.Label(window, text="파일 이름",width=100)
label3.place(x=150,y=150)

write_frame = ttk.Frame(window)

label4 = ttk.Label(write_frame, text="입력을 시작할 행번호   : ")
label4.grid(row=0,column=1)

start_row=ttk.Entry(write_frame)
start_row.grid(row=0,column=2)

button4 = ttk.Button(write_frame, text="파일 출력", command = lambda:sheet_export())
button4.grid(row=0,column=3)

check_accepted2 = ttk.Label(write_frame, text="")
check_accepted2.grid(row=0,column=4)

write_frame.place(x=10,y=180)


warning = ttk.Label(window, text="※주의사항※\n-불러오는 파일의 Sheet1에 자료가 담겨있어야 합니다.\n-저장하는 파일의 시트이름은 '택배출고요청 발주서'로 저장되어있어야 합니다.")
warning.place(x=10,y=210)

copy_right = ttk.Label(window, text = "COPYRIGHT ⓒ 2020 GUNHEE ALL RIGHTS RESERVED.")
copy_right.place(x=650,y=260)

version = ttk.Label(window, text = "v0.0.2")
version.place(x=910,y=280)
window.mainloop()